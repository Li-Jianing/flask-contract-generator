from flask import Flask, request, render_template_string, send_file, redirect, url_for, session
import os
import re
import uuid
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'contract_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['GENERATED_FOLDER'] = 'generated'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['GENERATED_FOLDER'], exist_ok=True)

# 从 .xlsx 中提取所有 {{字段名}}
def extract_placeholders_xlsx(xlsx_path):
    wb = load_workbook(xlsx_path)
    fields = set()
    pattern = r'\{\{(.*?)\}\}'
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    fields.update(re.findall(pattern, cell.value))
    return list(fields)

# 替换所有 {{字段}}，生成新 Excel
def fill_template_xlsx(template_path, data):
    wb = load_workbook(template_path)
    pattern = r'\{\{(.*?)\}\}'
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    new_val = cell.value
                    for match in re.findall(pattern, cell.value):
                        if match in data:
                            new_val = new_val.replace(f'{{{{{match}}}}}', data[match])
                    cell.value = new_val
    return wb

# 首页上传
@app.route('/', methods=['GET', 'POST'])
def upload_template():
    if request.method == 'POST':
        file = request.files['template']
        if file and file.filename.endswith('.xlsx'):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            fields = extract_placeholders_xlsx(filepath)
            session['template'] = filename
            session['fields'] = fields
            return redirect(url_for('fill_form'))
    return render_template_string('''
        <h2>上传Excel模板（.xlsx）</h2>
        <form method="post" enctype="multipart/form-data">
            <input type="file" name="template" accept=".xlsx" required>
            <input type="submit" value="上传">
        </form>
    ''')

# 动态生成表单
@app.route('/fill', methods=['GET', 'POST'])
def fill_form():
    template = session.get('template')
    fields = session.get('fields', [])
    if not template or not fields:
        return redirect(url_for('upload_template'))
    template_path = os.path.join(app.config['UPLOAD_FOLDER'], template)

    if request.method == 'POST':
        data = {field: request.form.get(field, '') for field in fields}
        wb = fill_template_xlsx(template_path, data)
        out_id = str(uuid.uuid4())
        out_xlsx = f'{out_id}.xlsx'
        out_path = os.path.join(app.config['GENERATED_FOLDER'], out_xlsx)
        wb.save(out_path)
        return render_template_string('''
            <h2>合同生成成功！</h2>
            <a href="{{ url_for('download_file', filename=filename) }}">下载Excel文档</a><br>
            <a href="/">返回首页</a>
        ''', filename=out_xlsx)

    form_html = '<h2>填写合同信息</h2><form method="post">'
    for field in fields:
        form_html += f'{field}：<input type="text" name="{field}" required><br>'
    form_html += '<input type="submit" value="生成合同"></form>'
    return render_template_string(form_html)

# 文件下载
@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['GENERATED_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return '文件不存在', 404

if __name__ == '__main__':
    app.run(debug=True)

